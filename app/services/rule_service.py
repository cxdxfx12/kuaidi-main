"""
运费规则管理服务
支持可视化配置、保存、读取
规则存储为 JSON 文件（data/config/fee_rules.json）
三层规则体系：客户级 → 区域级 → 全局级
"""
import os
import json
import math
from typing import Dict, List, Optional

from app.core.utils import parse_date

# 用于 Excel 解析/写入
import xlsxwriter
try:
    import openpyxl
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


def apply_weight_rounding(weight: float, mode: str, params: Optional[Dict] = None) -> float:
    """
    根据重量进位模式调整重量
    
    :param weight: 原始重量（kg）
    :param mode: 进位模式:
        - "actual": 实际重量
        - "round_05": 0.5进位（<0.5取0.5, ≥0.5取1）
        - "round_1": 四舍五入（<0.5舍去, ≥0.5取1）
        - "ceil_1kg": 向上取整（小数位都向上取整）
        - "segmented": 分段进位（自定义舍位/进位值）
        - "round_trunc": 进位舍位（只舍或只进）
    :param params: 进位参数:
        - segmented: {"segment_drop": 0.2, "segment_ceil": 0.7}
        - round_trunc: {"direction": "drop" | "ceil"}
    :return: 调整后的重量
    """
    params = params or {}
    
    if mode == "actual":
        return weight
    elif mode == "round_05":
        if weight <= 0.5:
            return 0.5
        return math.ceil(weight) if weight > 1.0 else 1.0
    elif mode == "round_1":
        return round(weight)
    elif mode == "ceil_1kg":
        return math.ceil(weight)
    elif mode == "segmented":
        segment_drop = params.get("segment_drop", 0.2)
        segment_ceil = params.get("segment_ceil", 0.7)
        if weight <= segment_drop:
            return 0.0
        elif weight <= segment_ceil:
            return (segment_drop + segment_ceil) / 2
        else:
            return math.ceil(weight) if weight > 1.0 else 1.0
    elif mode == "round_trunc":
        direction = params.get("direction", "drop")
        if direction == "drop":
            return float(int(weight))
        else:
            return math.ceil(weight)
    return weight


def calc_tiered_fee(weight: float,
                    tier_0_05: float, tier_05_1: float, tier_1_2: float, tier_2_3: float,
                    first_fee: float, continued_fee: float,
                    first_fee_30: float, continued_fee_30: float,
                    min_fee: float, continued_unit: str = "kg") -> float:
    """
    阶梯定价计算
    0-0.5kg → tier_0_05 (固定价)
    0.5-1kg → tier_05_1 (固定价)
    1-2kg   → tier_1_2  (固定价)
    2-3kg   → tier_2_3  (固定价)
    3-30kg  → 首重 first_fee (1kg) + 续重 continued_fee 每kg
    30kg+   → 首重 first_fee_30 (1kg) + 续重 continued_fee_30 每kg
    """
    if weight <= 0:
        weight = 0.01

    if weight <= 0.5:
        fee = tier_0_05
    elif weight <= 1.0:
        fee = tier_05_1
    elif weight <= 2.0:
        fee = tier_1_2
    elif weight <= 3.0:
        fee = tier_2_3
    elif weight < 30.0:
        continued_weight = weight - 1.0
        if continued_unit == "100g":
            units = math.ceil(continued_weight / 0.1)
            fee = first_fee + units * continued_fee
        else:
            fee = first_fee + continued_weight * continued_fee
    else:
        f30 = first_fee_30 if first_fee_30 > 0 else first_fee
        c30 = continued_fee_30 if continued_fee_30 > 0 else continued_fee
        continued_weight = weight - 1.0
        if continued_unit == "100g":
            units = math.ceil(continued_weight / 0.1)
            fee = f30 + units * c30
        else:
            fee = f30 + continued_weight * c30

    return round(max(fee, min_fee), 2)


class Rule:
    """单条计费规则"""
    def __init__(self, name: str = "", regions: str = "", stations: str = "",
                 min_weight: float = 0.0, max_weight: float = 999.0,
                 first_fee: float = 0.0, continued_fee: float = 0.0,
                 min_fee: float = 0.0, rule_type: str = "region",
                 continued_unit: str = "kg", weight_rounding: str = "actual",
                 rounding_params: Optional[Dict] = None,
                 计泡系数: float = 6000.0,
                 avg_weight_mode: bool = False, avg_weight_limit: float = 3.0,
                 pricing_mode: str = "standard",
                 tier_0_05: float = 0.0, tier_05_1: float = 0.0,
                 tier_1_2: float = 0.0, tier_2_3: float = 0.0,
                 first_fee_30: float = 0.0, continued_fee_30: float = 0.0,
                 avg_weight_deviation_step: float = 0.1,
                 avg_weight_deviation_surcharge: float = 0.0):
        """
        :param name: 规则名称
        :param regions: 逗号分隔的区域关键词，如 "上海,江苏,浙江"
        :param stations: 逗号分隔的网点编码，如 "ST001,ST002"
        :param rule_type: 规则类型: "station"(网点级), "region"(区域级), "global"(全局级)
        :param continued_unit: 续重单位: "kg"(全续), "100g"(百克续)
        :param weight_rounding: 重量进位模式: "actual", "round_05", "round_1", "ceil_1kg", "segmented", "round_trunc"
        :param rounding_params: 进位参数: {"segment_drop": 0.2, "segment_ceil": 0.7, "direction": "drop"}
        :param 计泡系数: 体积重除数，默认6000
        :param avg_weight_mode: 是否启用拉均重模式（同一客户+区域的所有包裹拉平均重量统一计费）
        :param avg_weight_limit: 均重上限(kg)，超出的包裹不参与均重，单独按实际重量计算
        :param pricing_mode: 计费模式: "standard"(标准首重+续重), "tiered"(阶梯定价)
        :param tier_0_05: 阶梯定价 0-0.5kg 固定价
        :param tier_05_1: 阶梯定价 0.5-1kg 固定价
        :param tier_1_2: 阶梯定价 1-2kg 固定价
        :param tier_2_3: 阶梯定价 2-3kg 固定价
        :param first_fee_30: 阶梯定价 30kg+ 首重费（3-30kg 段用 first_fee/continued_fee）
        :param continued_fee_30: 阶梯定价 30kg+ 续重费
        :param avg_weight_deviation_step: 均重偏差步长(kg)，实际重量每超出均重此步长即触发加价，默认0.1kg
        :param avg_weight_deviation_surcharge: 均重偏差加价(元)，每超出一个步长加价的金额，默认0(不启用)
        """
        self.name = name
        self.regions = regions
        self.stations = stations
        self.min_weight = float(min_weight)
        self.max_weight = float(max_weight)
        self.first_fee = float(first_fee)
        self.continued_fee = float(continued_fee)
        self.min_fee = float(min_fee)
        self.rule_type = rule_type  # "station", "region", "global"
        self.continued_unit = continued_unit  # "kg" or "100g"
        self.weight_rounding = weight_rounding  # 重量进位模式
        self.rounding_params = rounding_params or {}  # 进位参数
        self.计泡系数 = float(计泡系数)  # 体积重除数
        self.avg_weight_mode = bool(avg_weight_mode)  # 拉均重模式
        self.avg_weight_limit = float(avg_weight_limit)  # 均重上限
        self.pricing_mode = pricing_mode  # "standard" or "tiered"
        self.tier_0_05 = float(tier_0_05)
        self.tier_05_1 = float(tier_05_1)
        self.tier_1_2 = float(tier_1_2)
        self.tier_2_3 = float(tier_2_3)
        self.first_fee_30 = float(first_fee_30)
        self.continued_fee_30 = float(continued_fee_30)
        self.avg_weight_deviation_step = float(avg_weight_deviation_step)  # 均重偏差步长(kg)
        self.avg_weight_deviation_surcharge = float(avg_weight_deviation_surcharge)  # 偏差加价(元)

    def matches(self, region: str = "", station_code: str = "", weight: float = 0.0) -> bool:
        """判断是否匹配该规则"""
        # 网点匹配（支持中英文逗号）
        if self.stations.strip():
            _stations_norm = self.stations.replace("，", ",")
            station_list = [s.strip() for s in _stations_norm.split(",") if s.strip()]
            if station_code and station_code.strip() not in station_list:
                return False
        
        # 区域匹配（包含任一关键词即可）
        region_match = False
        region = (region or "").strip()
        if self.regions.strip():
            # 支持中英文逗号分割： "浙江,江苏" 或 "浙江，江苏"
            _regions_norm = self.regions.replace("，", ",")
            keywords = [k.strip() for k in _regions_norm.split(",") if k.strip()]
            region_match = any(k in region for k in keywords)
        else:
            region_match = True

        # 重量匹配
        weight_match = self.min_weight <= float(weight) <= self.max_weight

        return region_match and weight_match

    def to_dict(self) -> Dict:
        return {
            "name": self.name,
            "regions": self.regions,
            "stations": self.stations,
            "min_weight": self.min_weight,
            "max_weight": self.max_weight,
            "first_fee": self.first_fee,
            "continued_fee": self.continued_fee,
            "min_fee": self.min_fee,
            "rule_type": self.rule_type,
            "continued_unit": self.continued_unit,
            "weight_rounding": self.weight_rounding,
            "rounding_params": self.rounding_params,
            "计泡系数": self.计泡系数,
            "avg_weight_mode": self.avg_weight_mode,
            "avg_weight_limit": self.avg_weight_limit,
            "pricing_mode": self.pricing_mode,
            "tier_0_05": self.tier_0_05,
            "tier_05_1": self.tier_05_1,
            "tier_1_2": self.tier_1_2,
            "tier_2_3": self.tier_2_3,
            "first_fee_30": self.first_fee_30,
            "continued_fee_30": self.continued_fee_30,
            "avg_weight_deviation_step": self.avg_weight_deviation_step,
            "avg_weight_deviation_surcharge": self.avg_weight_deviation_surcharge,
        }

    @classmethod
    def from_dict(cls, data: Dict) -> "Rule":
        return cls(
            name=data.get("name", ""),
            regions=data.get("regions", ""),
            stations=data.get("stations", ""),
            min_weight=float(data.get("min_weight", 0)),
            max_weight=float(data.get("max_weight", 999)),
            first_fee=float(data.get("first_fee", 0)),
            continued_fee=float(data.get("continued_fee", 0)),
            min_fee=float(data.get("min_fee", 0)),
            rule_type=data.get("rule_type", "region"),
            continued_unit=data.get("continued_unit", "kg"),
            weight_rounding=data.get("weight_rounding", "actual"),
            rounding_params=data.get("rounding_params", {}),
            计泡系数=float(data.get("计泡系数", 6000)),
            avg_weight_mode=bool(data.get("avg_weight_mode", False)),
            avg_weight_limit=float(data.get("avg_weight_limit", 3.0)),
            pricing_mode=data.get("pricing_mode", "standard"),
            tier_0_05=float(data.get("tier_0_05", 0)),
            tier_05_1=float(data.get("tier_05_1", 0)),
            tier_1_2=float(data.get("tier_1_2", 0)),
            tier_2_3=float(data.get("tier_2_3", 0)),
            first_fee_30=float(data.get("first_fee_30", 0)),
            continued_fee_30=float(data.get("continued_fee_30", 0)),
            avg_weight_deviation_step=float(data.get("avg_weight_deviation_step", 0.1)),
            avg_weight_deviation_surcharge=float(data.get("avg_weight_deviation_surcharge", 0.0)),
        )


class RuleService:
    """规则管理服务 - 支持三层规则体系"""

    def __init__(self):
        from app.models.path_config import get_config_file
        self.config_file = get_config_file("fee_rules.json")
        self.default_settings_file = get_config_file("default_settings.json")
        # 加载"无重量默认价格"（取不到则用3.0兜底）
        self._empty_weight_fee = self._load_empty_weight_fee()

    def _load_empty_weight_fee(self) -> float:
        try:
            if os.path.exists(self.default_settings_file):
                with open(self.default_settings_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                val = data.get("empty_weight_fee")
                if val is not None:
                    return float(val)
        except Exception:
            pass
        return 3.0

    def load_rules(self) -> List[Rule]:
        """读取规则列表（按类型分组）"""
        if not os.path.exists(self.config_file):
            rules = self._create_default_rules()
            self.save_rules(rules)
            return rules

        try:
            with open(self.config_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            return [Rule.from_dict(r) for r in data.get("rules", [])]
        except Exception as e:
            print(f"读取规则失败: {e}")
            return self._create_default_rules()

    def _create_default_rules(self) -> List[Rule]:
        """创建默认规则（包含网点级示例规则）"""
        rules = []
        
        # 全局兜底规则（优先级最低）
        rules.append(Rule("全局规则", "", "", 0, 999, 6.0, 3.0, 6.0, "global"))
        
        # 区域级规则
        region_rules = [
            ("北京 - 华北地区", "北京", 4.0, 1.8, 4.0),
            ("天津 - 华北地区", "天津", 4.0, 1.8, 4.0),
            ("上海 - 华东地区", "上海", 3.5, 1.5, 3.5),
            ("重庆 - 西南地区", "重庆", 5.0, 2.5, 5.0),
            ("江苏 - 华东地区", "江苏", 3.5, 1.5, 3.5),
            ("浙江 - 华东地区", "浙江", 3.5, 1.5, 3.5),
            ("安徽 - 华东地区", "安徽", 3.5, 1.5, 3.5),
            ("福建 - 华东地区", "福建", 3.5, 1.5, 3.5),
            ("江西 - 华东地区", "江西", 3.5, 1.5, 3.5),
            ("山东 - 华东地区", "山东", 3.5, 1.5, 3.5),
            ("河北 - 华北地区", "河北", 4.0, 1.8, 4.0),
            ("山西 - 华北地区", "山西", 4.0, 1.8, 4.0),
            ("内蒙古 - 华北地区", "内蒙古", 4.0, 1.8, 4.0),
            ("广东 - 华南地区", "广东", 4.0, 1.8, 4.0),
            ("广西 - 华南地区", "广西", 4.0, 1.8, 4.0),
            ("海南 - 华南地区", "海南", 4.0, 1.8, 4.0),
            ("河南 - 华中地区", "河南", 4.5, 2.0, 4.5),
            ("湖北 - 华中地区", "湖北", 4.5, 2.0, 4.5),
            ("湖南 - 华中地区", "湖南", 4.5, 2.0, 4.5),
            ("四川 - 西南地区", "四川", 5.0, 2.5, 5.0),
            ("贵州 - 西南地区", "贵州", 5.0, 2.5, 5.0),
            ("云南 - 西南地区", "云南", 5.0, 2.5, 5.0),
            ("西藏 - 西南地区", "西藏", 5.0, 2.5, 5.0),
            ("陕西 - 西北地区", "陕西", 8.0, 4.0, 8.0),
            ("甘肃 - 西北地区", "甘肃", 8.0, 4.0, 8.0),
            ("青海 - 西北地区", "青海", 8.0, 4.0, 8.0),
            ("宁夏 - 西北地区", "宁夏", 8.0, 4.0, 8.0),
            ("新疆 - 西北地区", "新疆", 8.0, 4.0, 8.0),
            ("辽宁 - 东北地区", "辽宁", 5.0, 2.5, 5.0),
            ("吉林 - 东北地区", "吉林", 5.0, 2.5, 5.0),
            ("黑龙江 - 东北地区", "黑龙江", 5.0, 2.5, 5.0),
            ("香港 - 港澳台", "香港", 30.0, 20.0, 30.0),
            ("澳门 - 港澳台", "澳门", 30.0, 20.0, 30.0),
            ("台湾 - 港澳台", "台湾", 30.0, 20.0, 30.0),
        ]
        
        for name, regions, first_fee, continued_fee, min_fee in region_rules:
            rules.append(Rule(name, regions, "", 0, 999, first_fee, continued_fee, min_fee, "region"))
        
        # 网点级示例规则（优先级最高）
        rules.append(Rule("上海总部网点", "", "ST001", 0, 999, 3.0, 1.2, 3.0, "station"))
        rules.append(Rule("杭州旗舰网点", "", "ST002", 0, 999, 3.2, 1.3, 3.2, "station"))
        rules.append(Rule("北京核心网点", "", "ST003", 0, 999, 3.8, 1.6, 3.8, "station"))
        
        return rules

    def save_rules(self, rules: List[Rule], promotion_rules: Optional[List[Dict]] = None) -> bool:
        """保存规则列表（支持同时保存活动加价规则）"""
        try:
            # 清理旧版" - 华东地区/华北/西北"等大区后缀的默认规则（避免污染当前规则）
            cleaned_rules = []
            for r in rules:
                rname = r.name or ""
                # 旧默认区域名格式："省份 - 华东地区" / "省份 - 华北地区" / "省份 - 东北地区" / "省份 - 西北地区" / "省份 - 华南地区" / "省份 - 华中地区" / "省份 - 西南地区" / "省份 - 港澳台"
                legacy_suffixes = (
                    " - 华东地区", " - 华北地区", " - 东北地区", " - 西北地区",
                    " - 华南地区", " - 华中地区", " - 西南地区", " - 港澳台"
                )
                if any(rname.endswith(suf) for suf in legacy_suffixes):
                    # 跳过
                    continue
                cleaned_rules.append(r)

            data = {
                "version": "2.1",
                "description": "大圣运费计算规则 - 支持三层规则体系（客户级→区域级→全局级）+ 活动加价",
                "updated_at": __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "rules": [r.to_dict() for r in cleaned_rules],
                "promotion_rules": promotion_rules or []
            }
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存规则失败: {e}")
            return False

    # ============ Excel 导入/导出（批量导入客户+专属规则 ============

    PROVINCE_GROUP_MAP = {
        # key: group name in UI, value: list of province names
        "一区": ["浙江"],
        "二区": ["江苏", "安徽"],
        "三区": ["天津", "河北", "山东", "山西", "河南"],
        "四区": ["上海"],
        "五区": ["北京"],
        "六区": ["重庆"],
        "七区": ["广东", "广西", "海南", "福建", "江西", "湖南", "湖北"],
        "八区": ["黑龙江", "吉林", "辽宁", "内蒙古"],
        "九区": ["四川", "贵州", "云南"],
        "十区": ["甘肃", "宁夏", "青海", "陕西"],
        "十一区": ["新疆"],
        "十二区": ["西藏"],
        "十三区": ["海南"],  # 注：海南在原结构中存在
        "十四区": ["香港", "澳门", "台湾"],
    }

    # 反向映射：省份 → 组名
    _PROVINCE_TO_GROUP = {}
    for group, provinces in PROVINCE_GROUP_MAP.items():
        for p in provinces:
            _PROVINCE_TO_GROUP[p] = group
    del p
    del group
    del provinces

    WEIGHT_ROUNDING_TEXT_TO_MODE = {
        "实际重量": "actual",
        "0.5进位": "round_05",
        "四舍五入": "round_1",
        "向上取整": "ceil_1kg",
        "分段进位": "segmented",
        "进位舍位": "round_trunc",
    }
    CONTINUED_UNIT_TEXT_TO_MODE = {
        "全续": "kg",
        "百克续": "100g",
    }

    def generate_import_template(self, output_path: str) -> bool:
        """生成 Excel 导入模板（两个 Sheet：客户档案 + 客户专属规则）

        :param output_path: 输出路径（含 .xlsx 结尾）
        :return: 是否成功
        """
        try:
            wb = xlsxwriter.Workbook(output_path)

            # ---------- Sheet 1: 客户档案 ----------
            ws1 = wb.add_worksheet("客户档案")

            header_fmt = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
            required_fmt = wb.add_format({"italic": True, "color": "#C00000", "border": 1})
            normal_fmt = wb.add_format({"border": 1})

            headers_station = ["客户编码*", "客户名称*", "联系人", "联系电话", "地址", "是否启用", "备注"]
            for col, header in enumerate(headers_station):
                ws1.write(0, col, header, header_fmt)

            # 示例行：3 行
            sample_stations = [
                ["C001", "蜜丝婷", "张经理", "13800138000", "上海市浦东新区XX路XX号", 1, "大客户，每月结算"],
                ["C002", "珀莱雅", "李经理", "13800138001", "杭州市西湖区XX路XX号", 1, "大客户"],
                ["C010", "淘宝店铺A", "", "", "", 1, ""],
            ]
            for row_idx, row_data in enumerate(sample_stations):
                for col, value in enumerate(row_data):
                    ws1.write(row_idx + 1, col, value, normal_fmt)

            # 列宽
            ws1.set_column("A:A", 12)  # 客户编码
            ws1.set_column("B:B", 20)  # 客户名称
            ws1.set_column("C:C", 12)  # 联系人
            ws1.set_column("D:D", 16)  # 联系电话
            ws1.set_column("E:E", 40)  # 地址
            ws1.set_column("F:F", 10)  # 是否启用
            ws1.set_column("G:G", 30)  # 备注

            # ---------- Sheet 2: 客户专属规则 ----------
            ws2 = wb.add_worksheet("客户专属规则")

            headers_rule = [
                "客户编码*",
                "分区名称*",
                "定价模式",
                "0-0.5kg固定价",
                "0.5-1kg固定价",
                "1-2kg固定价",
                "2-3kg固定价",
                "首重(3-30kg)",
                "续重(3-30kg)*",
                "首重(30kg+)",
                "续重(30kg+)",
                "保底费(元)*",
                "续重单位",
                "重量进位",
                "拉均重",
                "均重上限",
                "偏差步长",
                "偏差加价",
            ]
            for col, header in enumerate(headers_rule):
                ws2.write(0, col, header, header_fmt)

            # 示例行：每行格式 [编码, 分区, 模式, 0-0.5, 0.5-1, 1-2, 2-3, 首重3-30, 续重3-30, 首30+, 续30+, 保底, 续重单位, 进位, 拉均重, 均重上限, 偏差步长, 偏差加价]
            # standard模式：阶梯列为0, 费用填在"首重(3-30kg)/续重(3-30kg)"中
            sample_rules = [
                ["C001", "一区", "standard", 0, 0, 0, 0, 3.5, 1.5, 0, 0, 2.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C001", "二区", "standard", 0, 0, 0, 0, 3.5, 1.5, 0, 0, 2.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C001", "三区", "standard", 0, 0, 0, 0, 4.0, 2.0, 0, 0, 3.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C001", "四区", "standard", 0, 0, 0, 0, 3.5, 1.5, 0, 0, 3.5, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C001", "五区", "standard", 0, 0, 0, 0, 4.0, 2.0, 0, 0, 3.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C001", "六区", "standard", 0, 0, 0, 0, 5.0, 2.5, 0, 0, 5.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C001", "七区", "standard", 0, 0, 0, 0, 4.5, 2.0, 0, 0, 4.5, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C001", "十四区", "standard", 0, 0, 0, 0, 30.0, 20.0, 0, 0, 30.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
                # C002: 珀莱雅 阶梯定价示例（tiered模式）
                ["C002", "一区", "tiered", 2.5, 3.5, 4.5, 6.0, 3.5, 1.5, 3.0, 1.0, 2.0, "全续", "实际重量", "是", 3.0, 0.1, 0.05],
                ["C002", "二区", "tiered", 2.5, 3.5, 4.5, 6.0, 3.5, 1.5, 3.0, 1.0, 2.0, "全续", "实际重量", "是", 3.0, 0.1, 0.05],
                # C010: 淘宝店铺A
                ["C010", "一区", "standard", 0, 0, 0, 0, 4.0, 2.0, 0, 0, 2.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
                ["C010", "二区", "standard", 0, 0, 0, 0, 4.0, 2.0, 0, 0, 2.0, "全续", "实际重量", "否", 3.0, 0.1, 0],
            ]

            for row_idx, row_data in enumerate(sample_rules):
                for col, value in enumerate(row_data):
                    ws2.write(row_idx + 1, col, value, normal_fmt)

            # 列宽
            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
                ws2.set_column(f"{col_letter}:{col_letter}", 12)
            ws2.set_column("A:A", 12)   # 客户编码
            ws2.set_column("B:B", 10)   # 分区名称
            ws2.set_column("C:C", 11)   # 定价模式
            ws2.set_column("O:O", 9)    # 拉均重
            ws2.set_column("P:P", 10)   # 均重上限
            ws2.set_column("Q:Q", 10)   # 偏差步长
            ws2.set_column("R:R", 10)   # 偏差加价

            wb.close()
            return True
        except Exception as e:
            return False

    def parse_import_excel(self, file_path: str) -> Dict:
        """解析导入 Excel 文件（使用 openpyxl）
        :return: {"stations": [...], "rules": [...], "errors": [...], "warnings": [...]}
        """
        try:
            result = {"stations": [], "rules": [], "errors": [], "warnings": []}

            if not _HAS_OPENPYXL:
                result["errors"].append("缺少 openpyxl 依赖包，请安装: pip install openpyxl")
                return result

            wb = openpyxl.load_workbook(file_path, data_only=True)

            # ---------- Sheet 1: 客户档案 ----------
            try:
                if "客户档案" not in wb.sheetnames:
                    result["warnings"].append("找不到 [客户档案] Sheet，跳过客户档案导入")
                else:
                    ws_stations = wb["客户档案"]
                    for row_idx, row in enumerate(ws_stations.iter_rows(min_row=2, values_only=True)):
                        if not row or row[0] is None or str(row[0]).strip() == "":
                            continue
                        code = str(row[0]).strip()
                        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                        contact = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                        phone = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                        address = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                        is_active_val = row[5] if len(row) > 5 and row[5] is not None else 1
                        try:
                            is_active = bool(int(float(is_active_val)))
                        except Exception:
                            is_active = True
                        remark = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""

                        if not name:
                            result["errors"].append(f"客户档案第 {row_idx + 2} 行（{code}）：客户名称不能为空")
                            continue
                        result["stations"].append({
                            "code": code, "name": name, "contact": contact,
                            "phone": phone, "address": address, "is_active": is_active,
                            "remark": remark, "row_num": row_idx + 2,
                        })
            except Exception as e:
                result["errors"].append(f"读取 [客户档案] Sheet 失败: {e}")

            # ---------- Sheet 2: 客户专属规则 ----------
            try:
                if "客户专属规则" not in wb.sheetnames:
                    result["warnings"].append("找不到 [客户专属规则] Sheet，跳过规则导入")
                else:
                    ws_rules = wb["客户专属规则"]
                    # 根据表头列数判断新旧格式：14列及以上=新格式(含阶梯定价/偏差加价)，7列=旧格式(仅标准定价)
                    header_row = next(ws_rules.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    is_new_format = header_row and len(header_row) >= 14

                    for row_idx, row in enumerate(ws_rules.iter_rows(min_row=2, values_only=True)):
                        if not row or row[0] is None or str(row[0]).strip() == "":
                            continue
                        if len(row) < 5:
                            continue

                        code = str(row[0]).strip()
                        group = str(row[1]).strip() if row[1] is not None else ""

                        if is_new_format:
                            # 新格式（18列）：编码, 分区, 定价模式, 0-0.5, 0.5-1, 1-2, 2-3, 首重3-30, 续重3-30, 首30+, 续30+, 保底, 续重单位, 进位, 拉均重, 均重上限, 偏差步长, 偏差加价
                            pricing_mode_text = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "standard"
                            tier_0_05_raw = row[3] if len(row) > 3 else 0
                            tier_05_1_raw = row[4] if len(row) > 4 else 0
                            tier_1_2_raw = row[5] if len(row) > 5 else 0
                            tier_2_3_raw = row[6] if len(row) > 6 else 0
                            first_fee_raw = row[7] if len(row) > 7 else 0
                            continued_fee_raw = row[8] if len(row) > 8 else 0
                            first_fee_30_raw = row[9] if len(row) > 9 else 0
                            continued_fee_30_raw = row[10] if len(row) > 10 else 0
                            min_fee_raw = row[11] if len(row) > 11 else 0
                            continued_unit_text = str(row[12]).strip() if len(row) > 12 and row[12] is not None else "全续"
                            weight_rounding_text = str(row[13]).strip() if len(row) > 13 and row[13] is not None else "实际重量"
                            avg_weight_mode_text = str(row[14]).strip() if len(row) > 14 and row[14] is not None else "否"
                            avg_weight_limit_raw = row[15] if len(row) > 15 else 3.0
                            deviation_step_raw = row[16] if len(row) > 16 else 0.1
                            deviation_surcharge_raw = row[17] if len(row) > 17 else 0.0
                        elif len(row) >= 14:
                            # 旧14列格式（无偏差加价）：编码, 分区, 定价模式, 0-0.5, 0.5-1, 1-2, 2-3, 首重3-30, 续重3-30, 首30+, 续30+, 保底, 续重单位, 进位
                            pricing_mode_text = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "standard"
                            tier_0_05_raw = row[3] if len(row) > 3 else 0
                            tier_05_1_raw = row[4] if len(row) > 4 else 0
                            tier_1_2_raw = row[5] if len(row) > 5 else 0
                            tier_2_3_raw = row[6] if len(row) > 6 else 0
                            first_fee_raw = row[7] if len(row) > 7 else 0
                            continued_fee_raw = row[8] if len(row) > 8 else 0
                            first_fee_30_raw = row[9] if len(row) > 9 else 0
                            continued_fee_30_raw = row[10] if len(row) > 10 else 0
                            min_fee_raw = row[11] if len(row) > 11 else 0
                            continued_unit_text = str(row[12]).strip() if len(row) > 12 and row[12] is not None else "全续"
                            weight_rounding_text = str(row[13]).strip() if len(row) > 13 and row[13] is not None else "实际重量"
                            avg_weight_mode_text = "否"
                            avg_weight_limit_raw = 3.0
                            deviation_step_raw = 0.1
                            deviation_surcharge_raw = 0.0
                        else:
                            # 旧格式（7列）：编码, 分区, 首重费, 续重费, 保底费, 续重单位, 进位
                            pricing_mode_text = "standard"
                            tier_0_05_raw = 0; tier_05_1_raw = 0; tier_1_2_raw = 0; tier_2_3_raw = 0
                            first_fee_30_raw = 0; continued_fee_30_raw = 0
                            first_fee_raw = row[2]
                            continued_fee_raw = row[3] if len(row) > 3 else 0
                            min_fee_raw = row[4] if len(row) > 4 else 0
                            continued_unit_text = str(row[5]).strip() if len(row) > 5 and row[5] is not None else "全续"
                            weight_rounding_text = str(row[6]).strip() if len(row) > 6 and row[6] is not None else "实际重量"
                            avg_weight_mode_text = "否"
                            avg_weight_limit_raw = 3.0
                            deviation_step_raw = 0.1
                            deviation_surcharge_raw = 0.0

                        if not group:
                            result["errors"].append(f"规则第 {row_idx + 2} 行：分区名称不能为空")
                            continue

                        try:
                            first_fee = float(first_fee_raw) if first_fee_raw is not None else 0.0
                            continued_fee = float(continued_fee_raw) if continued_fee_raw is not None else 0.0
                            min_fee = float(min_fee_raw) if min_fee_raw is not None else 0.0
                            tier_0_05 = float(tier_0_05_raw) if tier_0_05_raw is not None else 0.0
                            tier_05_1 = float(tier_05_1_raw) if tier_05_1_raw is not None else 0.0
                            tier_1_2 = float(tier_1_2_raw) if tier_1_2_raw is not None else 0.0
                            tier_2_3 = float(tier_2_3_raw) if tier_2_3_raw is not None else 0.0
                            first_fee_30 = float(first_fee_30_raw) if first_fee_30_raw is not None else 0.0
                            continued_fee_30 = float(continued_fee_30_raw) if continued_fee_30_raw is not None else 0.0
                            avg_weight_limit = float(avg_weight_limit_raw) if avg_weight_limit_raw is not None else 3.0
                            deviation_step = float(deviation_step_raw) if deviation_step_raw is not None else 0.1
                            deviation_surcharge = float(deviation_surcharge_raw) if deviation_surcharge_raw is not None else 0.0
                        except (ValueError, TypeError):
                            result["errors"].append(f"规则第 {row_idx + 2} 行（{code}/{group}）：费用字段必须是数字")
                            continue

                        pricing_mode = pricing_mode_text if pricing_mode_text in ("standard", "tiered") else "standard"
                        continued_unit = self.CONTINUED_UNIT_TEXT_TO_MODE.get(continued_unit_text, "kg")
                        weight_rounding = self.WEIGHT_ROUNDING_TEXT_TO_MODE.get(weight_rounding_text, "actual")

                        provinces = self._resolve_provinces_from_group(group)
                        if not provinces:
                            result["errors"].append(f"规则第 {row_idx + 2} 行（{code}/{group}）：未知的分区名或省份名")
                            continue

                        for province in provinces:
                            result["rules"].append({
                                "code": code, "province": province, "group": group,
                                "first_fee": first_fee, "continued_fee": continued_fee, "min_fee": min_fee,
                                "continued_unit": continued_unit, "weight_rounding": weight_rounding,
                                "pricing_mode": pricing_mode,
                                "tier_0_05": tier_0_05, "tier_05_1": tier_05_1,
                                "tier_1_2": tier_1_2, "tier_2_3": tier_2_3,
                                "first_fee_30": first_fee_30, "continued_fee_30": continued_fee_30,
                                "avg_weight_mode": avg_weight_mode_text in ("是", "yes", "true", "1"),
                                "avg_weight_limit": avg_weight_limit,
                                "deviation_step": deviation_step,
                                "deviation_surcharge": deviation_surcharge,
                                "row_num": row_idx + 2,
                            })
            except Exception as e:
                result["errors"].append(f"读取 [客户专属规则] Sheet 失败: {e}")

            wb.close()
            return result
        except Exception as e:
            return {
                "stations": [], "rules": [],
                "errors": [f"解析 Excel 文件失败: {e}"],
                "warnings": [],
            }

    def _resolve_provinces_from_group(self, group_text: str) -> List[str]:
        """从"分区名称"解析为实际省份列表
        :param group_text: 分区名（如"一区"、"北京"等
        :return: 省份列表，空 list 表示不合法
        """
        if not group_text:
            return []
        gt = group_text.strip()
        if gt in self.PROVINCE_GROUP_MAP:
            return list(self.PROVINCE_GROUP_MAP[gt])
        # 如果直接传入的是单个省份
        if gt in self._PROVINCE_TO_GROUP:
            return [gt]
        # 支持多省份，用逗号分隔？目前不支持，但简单判断
        return []

    def save_import_result(self, parsed_result: Dict, conflict_mode: str = "skip") -> Dict:
        """将解析结果写入数据库和 JSON 文件

        :param parsed_result: parse_import_excel() 的返回值
        :param conflict_mode: "overwrite"（覆盖）| "skip"（跳过已存在客户） | "append"（追加规则）
        :return: {"success": bool, "message": str, "stats": {...}}
        """
        try:
            stations = parsed_result.get("stations", [])
            rules = parsed_result.get("rules", [])

            if not stations and not rules:
                return {"success": False, "message": "Excel 文件为空或未解析出有效数据",
                        "stats": {"inserted_customers": 0, "inserted_rules": 0}}

            from app.models.database import get_session
            from app.models.station import Station
            session = get_session()

            stats = {
                "inserted_customers": 0, "updated_customers": 0, "skipped_customers": 0,
                "inserted_rules": 0, "updated_rules": 0, "skipped_rules": 0,
            }

            try:
                # ========= 1. 处理客户档案 =========
                existing_codes = {s.station_code for s in session.query(Station).all()}
                for s_data in stations:
                    code = s_data["code"]
                    if code in existing_codes:
                        if conflict_mode == "overwrite":
                            station = session.query(Station).filter(Station.station_code == code).first()
                            if station:
                                station.station_name = s_data["name"]
                                station.address = s_data.get("address", "")
                                station.contact_person = s_data.get("contact", "")
                                station.contact_phone = s_data.get("phone", "")
                                station.is_active = s_data.get("is_active", True)
                                stats["updated_customers"] += 1
                        else:
                            stats["skipped_customers"] += 1
                    else:
                        station = Station(station_code=code, station_name=s_data["name"],
                                          address=s_data.get("address", ""),
                                          contact_person=s_data.get("contact", ""),
                                          contact_phone=s_data.get("phone", ""),
                                          is_active=s_data.get("is_active", True))
                        session.add(station)
                        stats["inserted_customers"] += 1
                session.commit()

                # ========= 2. 处理客户规则 =========
                all_rules = self.load_rules()
                existing_rule_keys = set()
                for r in all_rules:
                    if r.rule_type == "station" and r.stations and r.stations.strip():
                        station_list = [s.strip() for s in r.stations.split(",") if s.strip()]
                        province_parts = (r.regions or "").split(",")
                        province_list = [p.strip() for p in province_parts if p and p.strip()]
                        for sc in station_list:
                            for pr in province_list:
                                existing_rule_keys.add((sc, pr.strip()))

                # 冲突处理：overwrite 模式下，先删除这些客户的旧规则
                codes_to_update = {r["code"] for r in rules}
                if conflict_mode == "overwrite":
                    filtered_rules = []
                    for r in all_rules:
                        if r.rule_type == "station" and r.stations and r.stations.strip():
                            station_list = [s.strip() for s in r.stations.split(",") if s.strip()]
                            if any(sc in codes_to_update for sc in station_list):
                                continue  # 删除这些规则，后面用新的替换
                        filtered_rules.append(r)
                    all_rules = filtered_rules

                # 添加新规则
                for r_data in rules:
                    key = (r_data["code"], r_data["province"])
                    if key in existing_rule_keys and conflict_mode != "overwrite":
                        stats["skipped_rules"] += 1
                        continue
                    new_rule = Rule(
                        name=f"{r_data['code']} - {r_data['group']}",
                        regions=r_data["province"], stations=r_data["code"],
                        min_weight=0.0, max_weight=999.0,
                        first_fee=float(r_data["first_fee"]),
                        continued_fee=float(r_data["continued_fee"]),
                        min_fee=float(r_data["min_fee"]),
                        rule_type="station",
                        continued_unit=r_data.get("continued_unit", "kg"),
                        weight_rounding=r_data.get("weight_rounding", "actual"),
                        rounding_params={},
                        pricing_mode=r_data.get("pricing_mode", "standard"),
                        tier_0_05=float(r_data.get("tier_0_05", 0)),
                        tier_05_1=float(r_data.get("tier_05_1", 0)),
                        tier_1_2=float(r_data.get("tier_1_2", 0)),
                        tier_2_3=float(r_data.get("tier_2_3", 0)),
                        first_fee_30=float(r_data.get("first_fee_30", 0)),
                        continued_fee_30=float(r_data.get("continued_fee_30", 0)),
                        avg_weight_mode=bool(r_data.get("avg_weight_mode", False)),
                        avg_weight_limit=float(r_data.get("avg_weight_limit", 3.0)),
                        avg_weight_deviation_step=float(r_data.get("deviation_step", 0.1)),
                        avg_weight_deviation_surcharge=float(r_data.get("deviation_surcharge", 0.0)),
                    )
                    all_rules.append(new_rule)
                    stats["inserted_rules"] += 1

                self.save_rules(all_rules)
                session.close()
                return {
                    "success": True,
                    "message": f"成功导入 {stats['inserted_customers']} 个客户，{stats['inserted_rules']} 条规则",
                    "stats": stats,
                }
            except Exception as e:
                session.rollback()
                session.close()
                return {"success": False, "message": f"保存失败: {e}", "stats": stats}
        except Exception as e:
            return {"success": False, "message": f"导入失败: {e}",
                    "stats": {"inserted_customers": 0, "inserted_rules": 0}}

    def load_promotion_rules(self) -> List[Dict]:
        """加载活动加价规则"""
        try:
            with open(self.config_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data.get("promotion_rules", [])
        except Exception:
            return []

    def calculate_fee(self, weight: float, region: str = "", station_code: str = "",
                      avg_weight: Optional[float] = None) -> Dict:
        """
        根据重量、区域、网点计算运费（三层规则匹配 + 活动加价）
        支持阶梯定价、拉均重、偏差加价
        :param weight: 重量(kg)
        :param region: 区域地址
        :param station_code: 客户编码
        :param avg_weight: 模拟均重(kg)，用于测试拉均重/偏差加价。传 None 则不启用均重逻辑
        :return: {"fee": 金额, "rule_name": 命中的规则名, "is_exception": 是否异常, "remark": 备注}
        """
        # 无重量：使用"无重量默认价格"（不再是异常）
        if weight is None or weight <= 0:
            return {
                "fee": self._empty_weight_fee,
                "rule_name": "无重量默认价",
                "is_exception": False,
                "remark": f"订单无重量，按默认价 ¥{self._empty_weight_fee:.2f} 结算"
            }

        rules = self.load_rules()

        # 第一步：查找客户专属规则（最优先）
        # 修复：stations 字段可能是逗号分隔值，必须按列表匹配；同时在循环内检查区域+重量
        station_rule = None
        sc = (station_code or "").strip()
        sc_list = [s.strip() for s in sc.split(",") if s.strip()] if sc else []
        for r in rules:
            if r.rule_type == "station" and r.stations and r.stations.strip():
                r_station_list = [s.strip() for s in r.stations.split(",") if s.strip()]
                customer_match = any(s in r_station_list for s in sc_list) or (not sc_list and False)
                if customer_match:
                    if r.first_fee > 0 or r.continued_fee > 0 or r.min_fee > 0 or r.regions:
                        # 在循环内直接检查区域+重量，找到完全匹配的客户规则
                        if r.matches(region, station_code, weight):
                            station_rule = r
                            break

        # 如果客户有专属规则且匹配，则使用
        if station_rule:
            matched_rule = station_rule
        else:
            # 第二步：匹配区域规则
            matched_rule = None
            for r in rules:
                if r.rule_type == "region" and r.matches(region, station_code, weight):
                    matched_rule = r
                    break

            # 第三步：使用全局规则兜底
            if not matched_rule:
                for r in rules:
                    if r.rule_type == "global" and r.matches(region, station_code, weight):
                        matched_rule = r
                        break

        if not matched_rule:
            return {
                "fee": 0.0,
                "rule_name": "无匹配规则",
                "is_exception": True,
                "remark": f"区域[{region}] 客户[{station_code}] 未匹配到计费规则"
            }

        # ===== 确定计费重量（支持拉均重模拟） =====
        billing_weight = weight
        avg_note = ""
        if avg_weight is not None and avg_weight > 0 and matched_rule.avg_weight_mode:
            billing_weight = avg_weight
            avg_note = f" [均重{avg_weight:.2f}kg]"

        # ===== 计算基础运费 =====
        first_fee = matched_rule.first_fee
        continued_fee = matched_rule.continued_fee
        continued_unit = matched_rule.continued_unit
        weight_rounding = matched_rule.weight_rounding
        rounding_params = matched_rule.rounding_params

        # 阶梯定价模式
        if matched_rule.pricing_mode == "tiered":
            fee = calc_tiered_fee(
                billing_weight,
                matched_rule.tier_0_05, matched_rule.tier_05_1,
                matched_rule.tier_1_2, matched_rule.tier_2_3,
                first_fee, continued_fee,
                matched_rule.first_fee_30, matched_rule.continued_fee_30,
                matched_rule.min_fee, continued_unit
            )
            calc_mode = "阶梯"
        else:
            # 标准模式：首重+续重
            rounded_weight = apply_weight_rounding(billing_weight, weight_rounding, rounding_params)

            if rounded_weight <= 1.0:
                fee = first_fee
            else:
                continued_weight = rounded_weight - 1.0
                if continued_unit == "100g":
                    units = math.ceil(continued_weight / 0.1)
                    fee = first_fee + units * continued_fee
                else:
                    fee = first_fee + continued_weight * continued_fee

            fee = round(max(fee, matched_rule.min_fee), 2)
            calc_mode = "标准"

        base_fee = fee

        # ===== 偏差加价（拉均重模式下，实际重量超出均重时触发） =====
        deviation_note = ""
        if (avg_weight is not None and avg_weight > 0
                and matched_rule.avg_weight_mode
                and matched_rule.avg_weight_deviation_surcharge > 0
                and weight > avg_weight):
            deviation_step = matched_rule.avg_weight_deviation_step
            deviation_surcharge = matched_rule.avg_weight_deviation_surcharge
            exceeded = weight - avg_weight
            steps = math.ceil(round(exceeded / deviation_step, 10))
            surcharge = round(steps * deviation_surcharge, 2)
            if surcharge > 0:
                fee = round(fee + surcharge, 2)
                deviation_note = f" 偏差+{steps}步×¥{deviation_surcharge}=+¥{surcharge}"

        # ===== 活动加价 =====

        # 第四步：应用活动加价（使用统一的日期解析，支持多种格式）
        promo_name = ""
        promo_amount = 0.0
        try:
            from datetime import datetime
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

            promo_rules = self.load_promotion_rules()
            for pr in promo_rules:
                try:
                    start = parse_date(pr.get("start_date", ""))
                    end = parse_date(pr.get("end_date", ""))
                    if start is None or end is None:
                        continue
                    if not (start <= today <= end):
                        continue

                    # 省份限定检查（regions留空=不限定）
                    regions_val = str(pr.get("regions", "")).strip()
                    if regions_val and region:
                        region_kws = [k.strip() for k in regions_val.split(",") if k.strip()]
                        if region_kws and not any(k in region for k in region_kws):
                            continue

                    markup_type = str(pr.get("markup_type", "percent")).strip().lower()
                    try:
                        markup_value = float(str(pr.get("markup_value", "0")).strip())
                    except (ValueError, TypeError):
                        continue

                    if markup_value <= 0:
                        continue

                    if markup_type == "fixed":
                        promo_amount = markup_value
                    elif markup_type == "weight":
                        promo_amount = weight * markup_value
                    elif markup_type == "percent":
                        promo_amount = base_fee * (markup_value / 100.0)
                    else:
                        continue

                    promo_name = str(pr.get("name", "活动加价"))
                    if regions_val:
                        promo_name = f"{promo_name}[{regions_val}]"
                    promo_amount = round(promo_amount, 2)
                    break
                except Exception:
                    continue
        except Exception:
            pass

        # 构建备注
        remark_parts = []
        remark_parts.append(f"{calc_mode}计费")
        if avg_note:
            remark_parts.append(avg_note.strip())
        remark_parts.append(f"基础¥{base_fee:.2f}")
        if deviation_note:
            remark_parts.append(deviation_note.strip())
        remark_parts.append(f"规则类型: {matched_rule.rule_type}")

        rule_label = matched_rule.name
        if avg_note:
            rule_label += avg_note
        if deviation_note:
            rule_label += deviation_note

        if promo_amount > 0:
            fee = round(fee + promo_amount, 2)
            return {
                "fee": fee,
                "rule_name": f"{rule_label} + {promo_name}(+¥{promo_amount})",
                "is_exception": False,
                "remark": " | ".join(remark_parts) + f" | 活动+¥{promo_amount}"
            }

        return {
            "fee": fee,
            "rule_name": rule_label,
            "is_exception": False,
            "remark": " | ".join(remark_parts)
        }

    def get_rules_by_type(self, rule_type: str) -> List[Rule]:
        """按类型获取规则"""
        rules = self.load_rules()
        return [r for r in rules if r.rule_type == rule_type]

    def add_station_rule(self, station_code: str, station_name: str, 
                         first_fee: float, continued_fee: float, min_fee: float):
        """快速添加网点专属规则"""
        rules = self.load_rules()
        new_rule = Rule(
            name=f"{station_name} - 网点专属",
            regions="",
            stations=station_code,
            first_fee=first_fee,
            continued_fee=continued_fee,
            min_fee=min_fee,
            rule_type="station"
        )
        rules.append(new_rule)
        return self.save_rules(rules)

    def copy_region_rules_to_stations(self, station_codes: List[str], region_name: str):
        """将区域规则批量应用到多个网点"""
        rules = self.load_rules()
        
        region_rule = None
        for r in rules:
            if r.rule_type == "region" and region_name in r.name:
                region_rule = r
                break
        
        if not region_rule:
            return False, f"未找到区域规则: {region_name}"
        
        for code in station_codes:
            new_rule = Rule(
                name=f"网点{code} - 继承{region_name}",
                regions="",
                stations=code,
                min_weight=region_rule.min_weight,
                max_weight=region_rule.max_weight,
                first_fee=region_rule.first_fee,
                continued_fee=region_rule.continued_fee,
                min_fee=region_rule.min_fee,
                rule_type="station"
            )
            rules.append(new_rule)
        
        success = self.save_rules(rules)
        return success, f"已为 {len(station_codes)} 个网点创建继承规则"